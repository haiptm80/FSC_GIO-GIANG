"""
FSC_AI AGENT GIO GIANG — FPT QA AGENT (V9.0).

Owner: Phạm Thị Minh Hải — MSNV: 00234640 (FE HN / QA-KSCL).
Tiêu chuẩn: ISO 21001:2018 | Khung KWSR Pipeline.
"""

import io
import json
import os
import re
from datetime import datetime

import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 1. THIẾT LẬP CẤU HÌNH TRANG STREAMLIT
# ==========================================
st.set_page_config(
    page_title="FSC AI AGENT GIO GIANG — FPT QA Agent V9.0",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded",
)

CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* ==========================================
   TIÊU ĐỀ CHÍNH & PHÂN ĐOẠN: MÀU XANH DƯƠNG ĐẬM (#0B2265), CHỮ BÉO, CỠ CHỮ 20px - 24px
   ========================================== */
h1, .fpt-title-h1, div[data-testid="stMarkdownContainer"] h1 {
    color: #0B2265 !important;
    font-size: 24px !important;
    font-weight: 900 !important;
    letter-spacing: -0.025em;
    margin-bottom: 0.75rem !important;
}

h2, .fpt-title-h2, div[data-testid="stMarkdownContainer"] h2 {
    color: #0B2265 !important;
    font-size: 22px !important;
    font-weight: 850 !important;
    letter-spacing: -0.02em;
    margin-top: 1rem !important;
    margin-bottom: 0.5rem !important;
}

h3, .fpt-title-h3, div[data-testid="stMarkdownContainer"] h3 {
    color: #0B2265 !important;
    font-size: 20px !important;
    font-weight: 850 !important;
    margin-top: 0.75rem !important;
    margin-bottom: 0.4rem !important;
}

h4, .fpt-title-h4, div[data-testid="stMarkdownContainer"] h4 {
    color: #0B2265 !important;
    font-size: 18px !important;
    font-weight: 800 !important;
    margin-top: 0.5rem !important;
    margin-bottom: 0.3rem !important;
}

/* ==========================================
   TABS CHÍNH TRÊN CÙNG (1-5): CHỮ MÀU XANH DƯƠNG, BÉO (BOLD 800), CỠ CHỮ 20px
   ========================================== */
div[data-testid="stAppViewContainer"] > section > div > div > div > div > div[data-baseweb="tab-list"] button[data-baseweb="tab"],
div[data-testid="stAppViewContainer"] > section > div > div > div > div > div[data-baseweb="tab-list"] button[data-baseweb="tab"] *,
div[data-testid="stTabs"]:first-child > div[data-baseweb="tab-list"] button[data-baseweb="tab"],
div[data-testid="stTabs"]:first-child > div[data-baseweb="tab-list"] button[data-baseweb="tab"] * {
    font-size: 20px !important;
    font-weight: 850 !important;
    color: #0B2265 !important; /* Xanh dương đậm */
}

div[data-testid="stTabs"]:first-child > div[data-baseweb="tab-list"] button[data-baseweb="tab"][aria-selected="true"],
div[data-testid="stTabs"]:first-child > div[data-baseweb="tab-list"] button[data-baseweb="tab"][aria-selected="true"] * {
    color: #0B2265 !important;
    border-bottom: 4px solid #0B2265 !important;
    font-weight: 900 !important;
}

/* ==========================================
   TABS PHÂN LOẠI 4 BẢNG (SUB-TABS): CHỮ MÀU XANH DƯƠNG, BÉO (BOLD 800), CỠ CHỮ 14px
   ========================================== */
div[data-testid="stTabPanel"] [role="tab"],
div[data-testid="stTabPanel"] [role="tab"] *,
div[data-testid="stTabPanel"] button[data-baseweb="tab"],
div[data-testid="stTabPanel"] button[data-baseweb="tab"] *,
div[data-testid="stTabContent"] [role="tab"],
div[data-testid="stTabContent"] [role="tab"] * {
    font-size: 14px !important;
    font-weight: 800 !important;
    color: #0B2265 !important;
}

div[data-testid="stTabPanel"] [role="tab"][aria-selected="true"],
div[data-testid="stTabPanel"] [role="tab"][aria-selected="true"] *,
div[data-testid="stTabContent"] [role="tab"][aria-selected="true"],
div[data-testid="stTabContent"] [role="tab"][aria-selected="true"] * {
    color: #0B2265 !important;
    border-bottom: 3.5px solid #0B2265 !important;
    font-weight: 900 !important;
}

/* ==========================================
   TIÊU ĐỀ TRONG BẢNG BIỂU (COLUMN HEADERS): CHỮ XANH DƯƠNG, BÉO, CỠ CHỮ 12px
   ========================================== */
div[data-testid="stDataFrame"] table th,
div[data-testid="stDataFrame"] th,
div[data-testid="stTable"] th,
div[data-testid="stDataFrame"] [class*="header"],
div[data-testid="stDataFrame"] div[role="columnheader"],
div[data-testid="stDataFrame"] div[role="columnheader"] *,
div[data-testid="stDataFrame"] .gdg-header,
div[data-testid="stDataFrameGlideDataEditor"] table th,
div[data-testid="stTable"] thead tr th {
    color: #0B2265 !important;
    font-weight: 800 !important;
    font-size: 12px !important;
}

/* ==========================================
   BẢNG DỮ LIỆU: VIỀN NỔI BẬT & ĐỔ MÀU XANH LÁ CÂY NHẸ
   ========================================== */
[data-testid="stDataFrame"], div[data-testid="stTable"] {
    border: 2.5px solid #10b981 !important;
    border-radius: 12px !important;
    background-color: #f0fdf4 !important;
    box-shadow: 0 4px 15px rgba(16, 185, 129, 0.15) !important;
    padding: 8px !important;
}

[data-testid="stDataFrame"] > div {
    background-color: #f0fdf4 !important;
    border-radius: 8px !important;
}

.table-highlight-box {
    background-color: #f0fdf4 !important;
    border: 2px solid #10b981 !important;
    border-radius: 12px !important;
    padding: 12px !important;
    margin-bottom: 1rem !important;
    box-shadow: 0 4px 12px rgba(16, 185, 129, 0.1) !important;
}

/* Header Container */
.fpt-header {
    background: linear-gradient(135deg, #0B2265 0%, #1e3a8a 50%, #05133d 100%);
    color: white;
    padding: 1.35rem 1.85rem;
    border-radius: 1rem;
    border-bottom: 4px solid #F26F21;
    box-shadow: 0 10px 25px -5px rgba(11, 34, 101, 0.25);
    margin-bottom: 1.5rem;
}

.fpt-badge {
    background-color: #F26F21;
    color: white;
    padding: 0.25rem 0.75rem;
    border-radius: 9999px;
    font-size: 0.8rem;
    font-weight: 800;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}

.iso-badge {
    background: rgba(255, 255, 255, 0.15);
    border: 1px solid rgba(255, 255, 255, 0.25);
    color: #fff;
    padding: 0.3rem 0.85rem;
    border-radius: 0.5rem;
    font-size: 0.85rem;
    font-weight: 700;
}

.metric-card {
    background: white;
    padding: 1.25rem;
    border-radius: 0.75rem;
    border: 1.5px solid #cbd5e1;
    box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    border-top: 4px solid #0B2265;
    transition: transform 0.2s ease, box-shadow 0.2s ease;
}
.metric-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 6px 14px rgba(0,0,0,0.08);
    border-top-color: #F26F21;
}
.metric-val {
    font-size: 1.85rem;
    font-weight: 900;
    color: #0B2265;
    font-family: 'JetBrains Mono', monospace;
}
.metric-label {
    font-size: 0.85rem;
    font-weight: 750;
    color: #475569;
    text-transform: uppercase;
    letter-spacing: 0.025em;
}

.alert-info-box {
    background-color: #eff6ff;
    border-left: 4.5px solid #0B2265;
    padding: 0.95rem 1.35rem;
    border-radius: 0.5rem;
    color: #1e3a8a;
    font-size: 0.9rem;
}

.step-badge {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 26px;
    height: 26px;
    border-radius: 50%;
    background-color: #0B2265;
    color: white;
    font-size: 13px;
    font-weight: 800;
    margin-right: 8px;
}

div.stButton > button:first-child {
    background-color: #F26F21;
    color: white;
    border: none;
    border-radius: 0.5rem;
    font-weight: 700;
    padding: 0.55rem 1.35rem;
    transition: all 0.2s ease;
}
div.stButton > button:first-child:hover {
    background-color: #d95a12;
    color: white;
    border: none;
    box-shadow: 0 4px 12px rgba(242, 111, 33, 0.3);
}

code, pre {
    font-family: 'JetBrains Mono', monospace !important;
}
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# ==========================================
# 2. KHỞI TẠO SESSION STATE & CẤU HÌNH
# ==========================================
DEFAULT_CONFIG = {
    "k_th": 0.583333333,
    "k_thcs": 0.75,
    "k_hsg": 1.20,
    "k_pd": 1.00,
    "k_clb": 1.25,
    "k_sukien": 1.00,
    "exclusions": [
        "SHL",
        "SINH HOẠT LỚP",
        "TH",
        "TỰ HỌC",
        "CHÀO CỜ",
        "TRỰC BAN",
        "SINH HOẠT CHI ĐOÀN",
        "ĐẠI HỘI",
        "KHAI GIẢNG",
        "BẾ GIẢNG",
    ],
    "regex_hsg_lop": r"HSG|HỌC SINH GIỎI",
    "regex_pd_lop": r"PĐ|PD|PHỤ ĐẠO|ĐỘI TUYỂN|ĐT",
    "regex_clb_lop": r"CLB|CÂU LẠC BỘ",
    "regex_sukien_lop": r"SỰ KIỆN|EVENT",
    "custom_categories": [],
    "quota_threshold": 110.0,
    "owner": "Phạm Thị Minh Hải - 00234640",
    "unit": "FE HN / QA-KSCL",
}

if "config" not in st.session_state:
    st.session_state.config = DEFAULT_CONFIG.copy()

if "custom_categories" not in st.session_state.config:
    st.session_state.config["custom_categories"] = []

WELCOME_MSG = f"""Xin chào Quý Thầy/Cô và Cán bộ QA! 👋

Tôi là **Trợ lý AI Kiểm Soát Giờ Giảng FSC (V9.0)** (Owner: **{DEFAULT_CONFIG["owner"]}**).

💡 **Các tính năng tôi có thể hỗ trợ ngay:**
1. **Nạp & Xử lý:** Nạp file FSP và Đối soát lịch đổi tiết 3 chiều `[Ngày + Tiết + Lớp]`.
2. **Chuẩn hóa Account:** Tự động lấy **Account GV** chuẩn và bóc tách đuôi email `@fpt.edu.vn`.
3. **Quy tắc phân loại:** Dựa trên **Cột Lớp** (Môn *Dự án PDP* tại lớp thường là *Giờ Chính Khóa*).
4. **Danh mục tùy chỉnh:** Cấu hình các loại tiết riêng của trường (Custom Categories).
5. **Cô lập Quota 110h:** Tự động cô lập giờ TH + THCS/THPT và cảnh báo vượt định mức.
6. **Lệnh nhanh (Slash Commands):** `/validate`, `/standardize`, `/audit`, `/synthesize`, `/report`.
"""

if "messages" not in st.session_state:
    st.session_state.messages = [{"role": "assistant", "content": WELCOME_MSG}]

if "pipeline_results" not in st.session_state:
    st.session_state.pipeline_results = None

if "execution_logs" not in st.session_state:
    st.session_state.execution_logs = []


def get_current_time_str(fmt: str = "%Y-%m-%d %H:%M:%S") -> str:
    """Lấy thời gian hiện tại có định dạng."""
    return datetime.now().astimezone().strftime(fmt)


def log_event(msg: str, level: str = "INFO") -> str:
    """Ghi nhật ký tiến trình vào session state."""
    now = get_current_time_str("%H:%M:%S")
    symbols = {"INFO": "ℹ️", "SUCCESS": "✅", "WARN": "⚠️", "ERROR": "❌"}
    formatted_log = f"[{now}] {symbols.get(level, '•')} {msg}"
    if hasattr(st, "session_state") and "execution_logs" in st.session_state:
        st.session_state.execution_logs.append(formatted_log)
    return formatted_log


# ==========================================
# 3. TIỆN ÍCH XỬ LÝ & PIPELINE
# ==========================================
def clean_account_name(raw_val) -> str:
    """Loại bỏ phần đuôi email @FPT.EDU.VN để lấy đúng Account GV chuẩn."""
    if not raw_val or pd.isna(raw_val):
        return ""
    val_str = str(raw_val).strip()
    return re.sub(r"@.*$", "", val_str, flags=re.IGNORECASE).strip().upper()


def check_exclusion_match(text: str, keywords: list) -> tuple:
    """Kiểm tra từ khóa loại trừ thông minh."""
    if not text:
        return False, ""
    text_upper = str(text).upper().strip()

    if re.match(r"^TH\d+", text_upper):
        return True, "TỰ HỌC (TH)"

    for kw in keywords:
        kw_clean = kw.strip().upper()
        if not kw_clean:
            continue
        if len(kw_clean) <= 3:
            pattern = (
                r"(?:\b|^|\s|[_\-\.])" + re.escape(kw_clean) + r"(?:\b|$|\s|[_\-\.])"
            )
            if re.search(pattern, text_upper):
                return True, kw_clean
        elif kw_clean in text_upper:
            return True, kw_clean
    return False, ""


def smart_read_excel_df(file_or_path) -> tuple:
    """Đọc file Excel tự động quét 15 dòng đầu tìm header."""
    try:
        if isinstance(file_or_path, str):
            df_raw = pd.read_excel(file_or_path, header=None, nrows=15)
            header_idx = 0
            for idx, row in df_raw.iterrows():
                row_str = " ".join(
                    [str(val).upper() for val in row.values if pd.notna(val)]
                )
                if (
                    ("NGÀY" in row_str or "NGAY" in row_str)
                    and ("TIẾT" in row_str or "TIET" in row_str)
                    and ("LỚP" in row_str or "LOP" in row_str)
                ):
                    header_idx = idx
                    break
            df = pd.read_excel(file_or_path, header=header_idx)
        else:
            file_or_path.seek(0)
            df_raw = pd.read_excel(file_or_path, header=None, nrows=15)
            header_idx = 0
            for idx, row in df_raw.iterrows():
                row_str = " ".join(
                    [str(val).upper() for val in row.values if pd.notna(val)]
                )
                if (
                    ("NGÀY" in row_str or "NGAY" in row_str)
                    and ("TIẾT" in row_str or "TIET" in row_str)
                    and ("LỚP" in row_str or "LOP" in row_str)
                ):
                    header_idx = idx
                    break
            file_or_path.seek(0)
            df = pd.read_excel(file_or_path, header=header_idx)

        df.columns = [str(c).strip() for c in df.columns]
        return df, header_idx
    except (ValueError, KeyError, OSError, RuntimeError) as e:
        raise ValueError(f"Không thể đọc file Excel: {e}") from e


def find_col_in_df(df: pd.DataFrame, possible_names: list):
    """Tìm tên cột khớp với danh sách gợi ý theo thứ tự ưu tiên."""
    for name in possible_names:
        for col in df.columns:
            clean_col = col.replace("\n", " ").strip().lower()
            if name.lower() == clean_col:
                return col
    # Fallback partial match
    for name in possible_names:
        for col in df.columns:
            clean_col = col.replace("\n", " ").strip().lower()
            if name.lower() in clean_col:
                return col
    return None


def run_fsc_pipeline(fsp_source, swap_source=None, config: dict | None = None) -> dict:
    """Tiến trình xử lý và đối soát 7 bước chuẩn V9.0."""
    if config is None:
        config = st.session_state.config

    st.session_state.execution_logs = []
    log_event("Khởi chạy Tiến trình Xử lý Giờ Giảng FSC V9.0...", "INFO")

    df_fsp, h_idx = smart_read_excel_df(fsp_source)
    log_event(
        f"Nạp file FSP thành công ({len(df_fsp)} dòng thô, header dòng {h_idx + 1}).",
        "SUCCESS",
    )

    col_lop = find_col_in_df(df_fsp, ["Lớp", "Lop"])
    col_mon = find_col_in_df(df_fsp, ["Môn học", "Mon hoc", "Môn"])

    # Ưu tiên cột Username / Account trước, nếu không có mới lấy cột Người dạy
    col_acc = find_col_in_df(df_fsp, ["Username", "Account", "Email"])
    if not col_acc:
        col_acc = find_col_in_df(df_fsp, ["Người dạy", "Nguoi day"])

    col_ngay = find_col_in_df(df_fsp, ["Ngày", "Ngay"])
    col_tiet = find_col_in_df(df_fsp, ["Tiết", "Tiet"])
    col_tenbai = find_col_in_df(df_fsp, ["Tên bài dạy", "Ten bài day", "Nhận xét lớp"])

    if not col_lop or not col_mon or not col_acc:
        raise ValueError(
            "File FSP thiếu cột bắt buộc (Lớp, Môn học, Account/Người dạy)."
        )

    exploded_rows = []
    for _, row in df_fsp.iterrows():
        raw_lop = str(row[col_lop]).strip() if pd.notna(row[col_lop]) else ""
        raw_mon = str(row[col_mon]).strip() if pd.notna(row[col_mon]) else ""
        raw_acc_val = row[col_acc]
        raw_acc = clean_account_name(raw_acc_val)
        raw_ngay = (
            str(row[col_ngay]).strip() if col_ngay and pd.notna(row[col_ngay]) else ""
        )
        raw_tiet = (
            str(row[col_tiet]).strip() if col_tiet and pd.notna(row[col_tiet]) else ""
        )
        raw_tenbai = (
            str(row[col_tenbai]).strip()
            if col_tenbai and pd.notna(row[col_tenbai])
            else ""
        )

        if not raw_lop or raw_lop.lower() == "nan":
            continue

        classes = [c.strip() for c in raw_lop.split(",") if c.strip()]
        for single_class in classes:
            exploded_rows.append(
                {
                    "Account": raw_acc,
                    "Lớp": single_class,
                    "Môn học": raw_mon,
                    "Ngày": raw_ngay,
                    "Tiết": raw_tiet,
                    "Tên bài dạy": raw_tenbai,
                }
            )

    df_exploded = pd.DataFrame(exploded_rows)
    log_event(
        f"Bóc tách chuỗi lớp ghép hoàn tất: {len(df_exploded)} dòng chi tiết (Account đã chuẩn hóa bỏ đuôi email).",
        "SUCCESS",
    )

    table1_rows = []
    table3_rows = []

    regex_hsg_lop = re.compile(config["regex_hsg_lop"], re.IGNORECASE)
    regex_pd_lop = re.compile(config["regex_pd_lop"], re.IGNORECASE)
    regex_clb_lop = re.compile(config["regex_clb_lop"], re.IGNORECASE)
    regex_sk_lop = re.compile(config["regex_sukien_lop"], re.IGNORECASE)

    custom_cats = config.get("custom_categories", [])

    for _, row in df_exploded.iterrows():
        mon_val = row["Môn học"]
        lop_val = row["Lớp"]

        is_ex_mon, kw_mon = check_exclusion_match(mon_val, config["exclusions"])
        is_ex_lop, kw_lop = check_exclusion_match(lop_val, config["exclusions"])
        is_excluded = is_ex_mon or is_ex_lop
        matched_kw = kw_mon or kw_lop

        if is_excluded:
            table3_rows.append(
                {
                    "Account GV": row["Account"],
                    "Lớp": row["Lớp"],
                    "Môn học": row["Môn học"],
                    "Ngày": row["Ngày"],
                    "Tiết": row["Tiết"],
                    "Từ khóa phát hiện": matched_kw,
                    "Lý do loại trừ": f"Trùng từ khóa không tính giờ [{matched_kw}]",
                }
            )
        else:
            lop_upper = lop_val.upper()
            tiet_th = 0
            tiet_thcs = 0
            tiet_hsg = 0
            tiet_pd = 0
            tiet_clb = 0
            tiet_sk = 0
            tiet_custom = 0
            gio_custom = 0.0
            custom_label = ""
            is_custom_quota = False

            # Kiểm tra khớp Danh mục tùy chỉnh (Custom Categories)
            matched_cust = False
            for c_cat in custom_cats:
                reg_p = c_cat.get("regex", "")
                if reg_p:
                    try:
                        if re.search(reg_p, lop_upper, re.IGNORECASE):
                            tiet_custom = 1
                            gio_custom = 1.0 * float(c_cat.get("k", 1.0))
                            custom_label = c_cat.get("label", "Tùy Chỉnh")
                            is_custom_quota = bool(c_cat.get("quota", False))
                            matched_cust = True
                            break
                    except re.error:
                        pass

            if not matched_cust:
                if regex_hsg_lop.search(lop_upper):
                    tiet_hsg = 1
                elif regex_pd_lop.search(lop_upper):
                    tiet_pd = 1
                elif regex_clb_lop.search(lop_upper):
                    tiet_clb = 1
                elif regex_sk_lop.search(lop_upper):
                    tiet_sk = 1
                else:
                    match = re.search(r"\d+", lop_upper)
                    if match:
                        grade = int(match.group(0))
                        if 1 <= grade <= 5:
                            tiet_th = 1
                        else:
                            tiet_thcs = 1
                    else:
                        tiet_thcs = 1

            gio_th = tiet_th * config["k_th"]
            gio_thcs = tiet_thcs * config["k_thcs"]
            gio_hsg = tiet_hsg * config["k_hsg"]
            gio_pd = tiet_pd * config["k_pd"]
            gio_clb = tiet_clb * config["k_clb"]
            gio_sk = tiet_sk * config["k_sukien"]
            tong_gio = (
                gio_th + gio_thcs + gio_hsg + gio_pd + gio_clb + gio_sk + gio_custom
            )

            table1_rows.append(
                {
                    "Account GV": row["Account"],
                    "Lớp": row["Lớp"],
                    "Môn học": row["Môn học"],
                    "Ngày": row["Ngày"],
                    "Tiết": row["Tiết"],
                    "Tiết TH": tiet_th,
                    "Tiết THCS/THPT": tiet_thcs,
                    "Tiết HSG": tiet_hsg,
                    "Tiết PD_ĐT": tiet_pd,
                    "Tiết CLB": tiet_clb,
                    "Tiết Sự Kiện": tiet_sk,
                    "Tiết Tùy Chỉnh": tiet_custom,
                    "Giờ TH": round(gio_th, 4),
                    "Giờ THCS/THPT": round(gio_thcs, 4),
                    "Giờ HSG": round(gio_hsg, 4),
                    "Giờ PD_ĐT": round(gio_pd, 4),
                    "Giờ CLB": round(gio_clb, 4),
                    "Giờ Sự Kiện": round(gio_sk, 4),
                    "Giờ Tùy Chỉnh": round(gio_custom, 4),
                    "Loại Tiết Tùy Chỉnh": custom_label,
                    "Custom_Quota_Eligible": is_custom_quota,
                    "Tổng Giờ Quy Đổi": round(tong_gio, 4),
                }
            )

    df_table1 = pd.DataFrame(table1_rows)
    df_table3 = pd.DataFrame(table3_rows)
    log_event(f"Tách {len(df_table3)} tiết loại trừ sang Bảng 3.", "SUCCESS")
    log_event(f"Phân loại {len(df_table1)} tiết sang Bảng 1 theo Cột Lớp.", "SUCCESS")

    table4_rows = []
    if swap_source is not None:
        try:
            df_swap, _ = smart_read_excel_df(swap_source)
            col_sw_ngay = find_col_in_df(df_swap, ["Ngày", "Ngay"])
            col_sw_tiet = find_col_in_df(df_swap, ["Tiết", "Tiet"])
            col_sw_lop = find_col_in_df(df_swap, ["Lớp", "Lop"])
            col_sw_goc = find_col_in_df(
                df_swap,
                ["Account GV theo TKB", "Account gốc", "GV theo TKB", "GV Gốc"],
            )
            col_sw_thay = find_col_in_df(
                df_swap,
                ["Account GV dạy thay", "Account thay", "GV dạy thay", "GV Thay"],
            )

            coord_map = {}
            for _, r in df_exploded.iterrows():
                key = f"{r['Ngày']}_{r['Tiết']}_{r['Lớp']}".upper()
                if key not in coord_map:
                    coord_map[key] = []
                coord_map[key].append(r["Account"])

            for _, sw_row in df_swap.iterrows():
                ngay = (
                    str(sw_row[col_sw_ngay]).strip()
                    if col_sw_ngay and pd.notna(sw_row[col_sw_ngay])
                    else ""
                )
                tiet = (
                    str(sw_row[col_sw_tiet]).strip()
                    if col_sw_tiet and pd.notna(sw_row[col_sw_tiet])
                    else ""
                )
                lop = (
                    str(sw_row[col_sw_lop]).strip()
                    if col_sw_lop and pd.notna(sw_row[col_sw_lop])
                    else ""
                )
                gv_goc = (
                    clean_account_name(sw_row[col_sw_goc])
                    if col_sw_goc and pd.notna(sw_row[col_sw_goc])
                    else ""
                )
                gv_thay = (
                    clean_account_name(sw_row[col_sw_thay])
                    if col_sw_thay and pd.notna(sw_row[col_sw_thay])
                    else ""
                )

                coord_key = f"{ngay}_{tiet}_{lop}".upper()
                matched_accs = coord_map.get(coord_key, [])

                if not matched_accs:
                    status = "🔴 LỖI ĐỎ"
                    flag = (
                        f"Không tìm thấy tọa độ. FSP chưa cập nhật GV thay [{gv_thay}]"
                    )
                elif gv_thay in matched_accs:
                    status = "🟢 HỢP LỆ XANH"
                    flag = f"Đã cập nhật đúng GV dạy thay [{gv_thay}]"
                elif gv_goc in matched_accs:
                    status = "🟡 CẢNH BÁO VÀNG"
                    flag = f"FSP chưa trừ GV gốc [{gv_goc}], chưa cập nhật GV dạy thay"
                else:
                    status = "🔴 LỖI ĐỎ"
                    flag = f"FSP ghi nhận account khác: [{', '.join(matched_accs)}]"

                table4_rows.append(
                    {
                        "Ngày": ngay,
                        "Tiết": tiet,
                        "Lớp": lop,
                        "GV Theo TKB": gv_goc,
                        "GV Dạy Thay": gv_thay,
                        "Trạng Thái": status,
                        "Chi Tiết Đối Soát": flag,
                    }
                )
            log_event(
                f"Đối soát {len(table4_rows)} lượt đổi tiết sang Bảng 4.", "SUCCESS"
            )
        except (ValueError, KeyError, OSError, RuntimeError) as e:
            log_event(f"Cảnh báo đọc file Đổi Tiết: {e}", "WARN")

    df_table4 = pd.DataFrame(table4_rows)

    table2_rows = []
    if not df_table1.empty:
        grouped = df_table1.groupby("Account GV")
        for acc, group in grouped:
            tiet_th = group["Tiết TH"].sum()
            tiet_thcs = group["Tiết THCS/THPT"].sum()
            tiet_hsg = group["Tiết HSG"].sum()
            tiet_pd = group["Tiết PD_ĐT"].sum()
            tiet_clb = group["Tiết CLB"].sum()
            tiet_sk = group["Tiết Sự Kiện"].sum()
            tiet_cust = (
                group["Tiết Tùy Chỉnh"].sum()
                if "Tiết Tùy Chỉnh" in group.columns
                else 0
            )
            tong_so_tiet = (
                tiet_th
                + tiet_thcs
                + tiet_hsg
                + tiet_pd
                + tiet_clb
                + tiet_sk
                + tiet_cust
            )

            gio_th = group["Giờ TH"].sum()
            gio_thcs = group["Giờ THCS/THPT"].sum()
            gio_hsg = group["Giờ HSG"].sum()
            gio_pd = group["Giờ PD_ĐT"].sum()
            gio_clb = group["Giờ CLB"].sum()
            gio_sk = group["Giờ Sự Kiện"].sum()
            gio_cust = (
                group["Giờ Tùy Chỉnh"].sum()
                if "Giờ Tùy Chỉnh" in group.columns
                else 0.0
            )

            gio_cust_quota = 0.0
            if (
                "Custom_Quota_Eligible" in group.columns
                and "Giờ Tùy Chỉnh" in group.columns
            ):
                gio_cust_quota = group[group["Custom_Quota_Eligible"].eq(True)][
                    "Giờ Tùy Chỉnh"
                ].sum()

            gio_dinh_muc = gio_th + gio_thcs + gio_cust_quota
            tong_gio_luong = (
                gio_th + gio_thcs + gio_hsg + gio_pd + gio_clb + gio_sk + gio_cust
            )
            is_over = bool(gio_dinh_muc > config["quota_threshold"])

            table2_rows.append(
                {
                    "Account Giáo Viên": acc,
                    "Số Tiết TH": int(tiet_th),
                    "Số Tiết THCS/THPT": int(tiet_thcs),
                    "Số Tiết HSG": int(tiet_hsg),
                    "Số Tiết PD_ĐT": int(tiet_pd),
                    "Số Tiết CLB": int(tiet_clb),
                    "Số Tiết Sự Kiện": int(tiet_sk),
                    "Số Tiết Tùy Chỉnh": int(tiet_cust),
                    "Tổng Số Tiết": int(tong_so_tiet),
                    "Số Giờ TH": round(gio_th, 2),
                    "Số Giờ THCS/THPT": round(gio_thcs, 2),
                    "Số Giờ HSG": round(gio_hsg, 2),
                    "Số Giờ PD_ĐT": round(gio_pd, 2),
                    "Số Giờ CLB": round(gio_clb, 2),
                    "Số Giờ Sự Kiện": round(gio_sk, 2),
                    "Số Giờ Tùy Chỉnh": round(gio_cust, 2),
                    "Tổng Giờ Trả Lương": round(tong_gio_luong, 2),
                    "Tổng Giờ Xét Định Mức": round(gio_dinh_muc, 2),
                    "Cảnh Báo Vượt Định Mức (>110h)": (
                        "🚨 VƯỢT ĐỊNH MỨC" if is_over else "✅ Hợp lệ"
                    ),
                    "canh_bao_vuot_gio": is_over,
                }
            )

    df_table2 = pd.DataFrame(table2_rows)
    if not df_table2.empty:
        df_table2 = df_table2.sort_values(by="Tổng Giờ Trả Lương", ascending=False)
    log_event(f"Tổng hợp Số Tiết & Số Giờ cho {len(df_table2)} giáo viên.", "SUCCESS")

    excel_buffer = io.BytesIO()
    df_table1_export = (
        df_table1.drop(columns=["Custom_Quota_Eligible"])
        if "Custom_Quota_Eligible" in df_table1.columns
        else df_table1
    )
    df_table2_export = (
        df_table2.drop(columns=["canh_bao_vuot_gio"])
        if "canh_bao_vuot_gio" in df_table2.columns
        else df_table2
    )

    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        df_table1_export.to_excel(writer, sheet_name="Bảng 1 - Chi Tiết", index=False)
        df_table2_export.to_excel(
            writer, sheet_name="Bảng 2 - Tổng Hợp Định Mức", index=False
        )
        df_table3.to_excel(writer, sheet_name="Bảng 3 - Loại Trừ", index=False)
        df_table4.to_excel(writer, sheet_name="Bảng 4 - Log Đổi Tiết", index=False)

    excel_buffer.seek(0)
    wb = openpyxl.load_workbook(excel_buffer)
    fill_main = PatternFill(start_color="F26F21", end_color="F26F21", fill_type="solid")
    fill_tiet = PatternFill(start_color="0B2265", end_color="0B2265", fill_type="solid")
    hdr_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            col_name = str(cell.value or "")
            cell.fill = (
                fill_tiet
                if "Số Tiết" in col_name or "Tổng Số Tiết" in col_name
                else fill_main
            )
            cell.font = hdr_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        sheet.row_dimensions[1].height = 28
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
        for col in sheet.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 13), 40)

    final_excel_bytes = io.BytesIO()
    wb.save(final_excel_bytes)
    final_excel_bytes.seek(0)

    os.makedirs("03_Outputs", exist_ok=True)
    today_str = get_current_time_str("%Y-%m-%d")
    output_excel_path = os.path.join(
        "03_Outputs",
        f"{today_str}_FPT_QA_BaoCao_GioGiang_00234640_Pham_Thi_Minh_Hai.xlsx",
    )
    with open(output_excel_path, "wb") as f_out:
        f_out.write(final_excel_bytes.getvalue())
    log_event(f"Đã lưu thành phẩm vào {output_excel_path}", "SUCCESS")

    return {
        "df_table1": df_table1,
        "df_table2": df_table2,
        "df_table3": df_table3,
        "df_table4": df_table4,
        "excel_bytes": final_excel_bytes.getvalue(),
        "excel_path": output_excel_path,
    }


# ==========================================
# 4. HÀM ĐỊNH DẠNG MÀU CỘT & SỐ HIỂN THỊ
# ==========================================
def format_fsc_number_str(val, is_table1=False, col_name=""):
    """Định dạng số chuẩn QA FSC:

    - Số 0 hoặc 0.0000 hiển thị là "0"
    - Số nguyên hiển thị dạng số nguyên tự nhiên (ví dụ: "6", "5", "1")
    - Bảng 1: Các tiết THCS/THPT và số giờ quy đổi hiển thị 2 chữ số thập phân (ví dụ: "0.75", "1.25", "0.58")
    - Bảng 2: Số giờ làm tròn 1 chữ số sau dấu phẩy (ví dụ: "76.5", "77.3")
    """
    if val is None or pd.isna(val) or val == "":
        return "0"
    try:
        f_val = float(val)
        if abs(f_val) < 1e-5:
            return "0"
        if f_val.is_integer():
            return str(int(f_val))

        # Bảng 1: Hiển thị 2 chữ số thập phân cho tiết THCS/THPT và số giờ quy đổi
        if is_table1:
            return f"{f_val:.2f}"

        # Bảng 2 và các bảng khác: 1 chữ số sau dấu phẩy
        return f"{round(f_val, 1):.1f}"
    except (ValueError, TypeError):
        return str(val)


def style_fsc_dataframe(df: pd.DataFrame, is_table1: bool = False):
    """Định dạng màu sắc các cột và số hiển thị theo yêu cầu QA FSC:

    - Tiêu đề cột: Màu xanh dương (#0B2265), béo (800), cỡ 12px
    - Bảng 1: Số giờ THCS/THPT lấy 2 chữ số thập phân (0.75)
    - Số 0.0000 hiển thị là 0
    - Cột Số Tiết: Màu Xanh dương nhạt (#e0f2fe)
    - Cột Số Giờ: Màu Xanh lá cây (#dcfce7)
    - Cột Tổng Giờ & Cảnh Báo: Màu Cam nhạt (#ffedd5)
    """
    if df.empty:
        return df

    df_clean = df.copy()

    # Chuẩn hóa định dạng số cho các cột Số Tiết và Số Giờ thành chuỗi chuẩn sạch
    non_numeric_cols = {
        "Ngày",
        "Lớp",
        "Môn học",
        "Account GV",
        "Account Giáo Viên",
        "Chi Tiết Đối Soát",
        "Từ khóa phát hiện",
        "Lý do loại trừ",
        "Loại Tiết Tùy Chỉnh",
        "GV Theo TKB",
        "GV Dạy Thay",
        "Trạng Thái",
        "Cảnh Báo Vượt Định Mức (>110h)",
    }
    for col in df_clean.columns:
        if col not in non_numeric_cols and (
            "Tiết" in col or "Giờ" in col or "Tổng" in col or "Xét" in col
        ):
            df_clean[col] = df_clean[col].apply(
                lambda x, c=col: format_fsc_number_str(
                    x, is_table1=is_table1, col_name=c
                )
            )

    # Xác định danh sách các cột theo nhóm màu
    tiet_cols = [
        c
        for c in df_clean.columns
        if ("Tiết" in c or c == "Tiết") and "Tổng" not in c and "Tổng Số Tiết" not in c
    ]
    gio_cols = [
        c
        for c in df_clean.columns
        if "Giờ" in c and "Tổng" not in c and "Xét" not in c and "Định Mức" not in c
    ]
    tong_cols = [
        c
        for c in df_clean.columns
        if "Tổng" in c
        or "Xét" in c
        or "Cảnh Báo" in c
        or "Tổng Số Tiết" in c
        or "Tổng Giờ" in c
    ]

    styler = df_clean.style

    # Đổ màu Xanh dương nhạt cho các cột Số Tiết
    if tiet_cols:
        styler = styler.set_properties(
            subset=tiet_cols,
            **{
                "background-color": "#e0f2fe",
                "color": "#0369a1",
                "font-weight": "600",
                "text-align": "center",
            },
        )

    # Đổ màu Xanh lá cây cho các cột Số Giờ
    if gio_cols:
        styler = styler.set_properties(
            subset=gio_cols,
            **{
                "background-color": "#dcfce7",
                "color": "#15803d",
                "font-weight": "600",
                "text-align": "center",
            },
        )

    # Đổ màu Cam nhạt cho các cột Tổng Giờ & Cảnh Báo
    if tong_cols:
        styler = styler.set_properties(
            subset=tong_cols,
            **{
                "background-color": "#ffedd5",
                "color": "#c2410c",
                "font-weight": "bold",
                "text-align": "center",
            },
        )

    # Thiết lập tiêu đề cột: Chữ màu xanh dương (#0B2265), béo (800), cỡ 12px
    styler = styler.set_table_styles(
        [
            {
                "selector": "th",
                "props": [
                    ("color", "#0B2265"),
                    ("font-weight", "800"),
                    ("font-size", "12px"),
                    ("text-align", "center"),
                    ("background-color", "#f8fafc"),
                ],
            },
            {
                "selector": "th.col_heading",
                "props": [
                    ("color", "#0B2265"),
                    ("font-weight", "800"),
                    ("font-size", "12px"),
                ],
            },
            {"selector": "td", "props": [("font-size", "12.5px")]},
        ]
    )

    return styler


# ==========================================
# 5. TRỢ LÝ HỘI THOẠI AI CHATBOT ENGINE
# ==========================================
def answer_user_query(query: str, results: dict, config: dict) -> str:
    """Engine xử lý câu hỏi tự nhiên và Slash Commands."""
    q = query.strip()
    q_upper = q.upper()

    if q.startswith("/"):
        cmd = q.split()[0].lower()
        if cmd == "/help":
            return """### 📚 Danh sách các Slash Commands khả dụng:
- `/validate`: Kiểm tra tính toàn vẹn và định dạng dữ liệu đầu vào.
- `/standardize`: Bóc tách lớp ghép và phân loại tiết loại trừ sang Bảng 3.
- `/audit`: Đối soát 3 chiều FSP vs Lịch đổi tiết sang Bảng 4.
- `/synthesize`: Tổng hợp số tiết, số giờ và kiểm tra quota 110h.
- `/report`: Hiển thị tóm tắt báo cáo chất lượng giờ giảng toàn diện.
- `/status`: Kiểm tra trạng thái dữ liệu và cấu hình hệ số hiện tại."""

        if cmd == "/status":
            has_data = "✅ Đã nạp và xử lý" if results else "⚠️ Chưa nạp dữ liệu"
            cust_count = len(config.get("custom_categories", []))
            return f"""### ⚙️ Trạng thái Hệ thống FSC QA Agent (V9.0):
- **Owner:** {config["owner"]} ({config["unit"]})
- **Trạng thái Dữ liệu:** {has_data}
- **Hệ số Tiểu học (K_TH):** `{config["k_th"]}` (35 phút/tiết)
- **Hệ số THCS/THPT (K_THCS):** `{config["k_thcs"]}` (45 phút/tiết)
- **Hệ số HSG:** `{config["k_hsg"]}` | **PĐ:** `{config["k_pd"]}` | **CLB:** `{config["k_clb"]}`
- **Danh mục Tùy chỉnh (Custom):** `{cust_count}` loại tiết riêng.
- **Ngưỡng Định Mức (Quota):** `{config["quota_threshold"]} giờ` (Chỉ tính giờ TH + THCS/THPT)."""

        if cmd in ["/validate", "/standardize", "/audit", "/synthesize", "/report"]:
            if not results:
                return "⚠️ **Chưa có dữ liệu!** Vui lòng sang Tab 3 nạp dữ liệu trước."

            t1 = len(results["df_table1"])
            t2 = len(results["df_table2"])
            t3 = len(results["df_table3"])
            t4 = len(results["df_table4"])
            over_count = len(
                results["df_table2"][results["df_table2"]["canh_bao_vuot_gio"].eq(True)]
            )

            if cmd == "/validate":
                return f"✅ **Kết quả Kiểm tra (/validate):** Bóc tách thành công `{t1 + t3}` lượt tiết."
            if cmd == "/standardize":
                return f"✅ **Kết quả Chuẩn hóa (/standardize):** Tách `{t3}` tiết sang Bảng 3, `{t1}` tiết sang Bảng 1."
            if cmd == "/audit":
                return f"✅ **Kết quả Đối soát (/audit):** Đã đối soát 3 chiều cho `{t4}` lượt đổi tiết."
            if cmd in ["/synthesize", "/report"]:
                sum_salary = results["df_table2"]["Tổng Giờ Trả Lương"].sum()
                sum_quota = results["df_table2"]["Tổng Giờ Xét Định Mức"].sum()
                return f"""### 📊 Báo Cáo Tổng Hợp Chất Lượng Giờ Giảng (/report)
- **Tổng số giáo viên giảng dạy:** `{t2}` GV.
- **Tổng số tiết hợp lệ:** `{results["df_table2"]["Tổng Số Tiết"].sum()}` tiết.
- **Tổng Quỹ Giờ Trả Lương:** `{sum_salary:.2f}` giờ.
- **Tổng Quỹ Giờ Xét Định Mức:** `{sum_quota:.2f}` giờ chính khóa.
- **Cảnh báo vượt định mức (>110h):** `{over_count}` giáo viên (Chỉ tính giờ TH + THCS/THPT).
- **Tiết bị loại trừ (Bảng 3):** `{t3}` tiết."""

    if not results:
        return "⚠️ Hệ thống chưa nạp dữ liệu. Hãy sang Tab 3 để tải file hoặc dùng dữ liệu mẫu."

    df_t1 = results["df_table1"]
    df_t2 = results["df_table2"]
    df_t3 = results["df_table3"]
    df_t4 = results["df_table4"]

    if any(k in q_upper for k in ["VƯỢT", "ĐỊNH MỨC", "QUOTA", "110", "OVER"]):
        over_teachers = df_t2[df_t2["canh_bao_vuot_gio"].eq(True)]
        if over_teachers.empty:
            return "✅ **Không có giáo viên nào vượt định mức 110 giờ** trong kỳ này!"

        md_table = "| STT | Account | Tiết TH | Tiết THCS | Giờ TH | Giờ THCS | Giờ Định Mức | Tổng Giờ Lương |\n|:---:|:---|:---:|:---:|:---:|:---:|:---:|:---:|\n"
        for i, (_, r) in enumerate(over_teachers.iterrows(), 1):
            md_table += f"| {i} | **{r['Account Giáo Viên']}** | {r['Số Tiết TH']} | {r['Số Tiết THCS/THPT']} | {r['Số Giờ TH']} | {r['Số Giờ THCS/THPT']} | **{r['Tổng Giờ Xét Định Mức']}** | {r['Tổng Giờ Trả Lương']} |\n"

        return f"### ⚠️ Phát hiện {len(over_teachers)} Giáo viên Vượt Định Mức (> 110 Giờ):\n\n{md_table}"

    matched_teacher = None
    for acc in df_t2["Account Giáo Viên"].unique():
        clean_acc = str(acc).upper().strip()
        if clean_acc in q_upper or clean_acc.replace(".", "") in q_upper:
            matched_teacher = acc
            break

    if matched_teacher:
        r2 = df_t2[df_t2["Account Giáo Viên"] == matched_teacher].iloc[0]
        r1_list = df_t1[df_t1["Account GV"] == matched_teacher]
        mon_list = ", ".join(r1_list["Môn học"].unique())
        lop_list = ", ".join(r1_list["Lớp"].unique())
        status_quota = (
            "🚨 **VƯỢT ĐỊNH MỨC**"
            if r2["canh_bao_vuot_gio"]
            else "✅ **Hợp lệ trong định mức**"
        )

        return f"""### 👤 Thông Tin Chi Tiết Giáo Viên: `{matched_teacher}`
- **Trạng Thái Định Mức:** {status_quota}
- **Tổng Số Tiết:** **{r2["Tổng Số Tiết"]} tiết** (TH: {r2["Số Tiết TH"]} - THCS/THPT: {r2["Số Tiết THCS/THPT"]} - HSG: {r2["Số Tiết HSG"]} - PĐ: {r2["Số Tiết PD_ĐT"]} - CLB: {r2["Số Tiết CLB"]})
- **Tổng Giờ Trả Lương:** **{r2["Tổng Giờ Trả Lương"]} giờ**
- **Tổng Giờ Xét Định Mức:** **{r2["Tổng Giờ Xét Định Mức"]} giờ**
- **Môn Dạy:** {mon_list}
- **Lớp Dạy:** {lop_list} ({len(r1_list)} lượt tiết)"""

    if any(
        k in q_upper
        for k in ["LOẠI TRỪ", "BẢNG 3", "SHL", "SINH HOẠT", "TỰ HỌC", "CHÀO CỜ"]
    ):
        return f"### 🚫 Thống Kê Tiết Loại Trừ (Bảng 3):\n- Tổng số: `{len(df_t3)}` tiết (SHL, Tự học TH, Chào cờ, Trực ban...)."

    if any(k in q_upper for k in ["ĐỔI TIẾT", "DẠY THAY", "BẢNG 4", "ĐỐI SOÁT"]):
        if df_t4.empty:
            return "ℹ️ Chưa nạp file Đổi Tiết để đối soát."
        red = len(df_t4[df_t4["Trạng Thái"].str.contains("LỖI ĐỎ", na=False)])
        yellow = len(df_t4[df_t4["Trạng Thái"].str.contains("CẢNH BÁO VÀNG", na=False)])
        green = len(df_t4[df_t4["Trạng Thái"].str.contains("HỢP LỆ XANH", na=False)])
        return f"### 🔄 Kết Quả Đối Soát Đổi Tiết (Bảng 4):\n- Tổng: `{len(df_t4)}` lượt.\n- 🟢 Hợp lệ: `{green}` | 🟡 Cảnh báo: `{yellow}` | 🔴 Lỗi đỏ: `{red}`."

    if "PDP" in q_upper or "DỰ ÁN" in q_upper:
        return "### 🎓 Quy Chuẩn Môn 'Dự án PDP':\nPhân loại hoàn toàn theo **Cột Lớp**. Học tại lớp thường (10A1, 11B2...) tính là **Tiết Chính Khóa THCS/THPT** ($K = 0.75$)."

    return f"Dữ liệu hiện tại: `{len(df_t2)}` GV, `{df_t2['Tổng Giờ Trả Lương'].sum():.2f}` giờ quy đổi, `{len(df_t2[df_t2['canh_bao_vuot_gio'].eq(True)])}` GV vượt định mức."


# ==========================================
# 6. SIDEBAR: HỒ SƠ & THAO TÁC NHANH
# ==========================================
with st.sidebar:
    st.markdown(
        """
    <div style="text-align: center; padding: 0.5rem 0 1rem 0;">
        <div style="display: inline-block; background: linear-gradient(135deg, #F26F21, #d95a12); color: white; border-radius: 12px; padding: 12px 16px; font-size: 24px;">
            🎓
        </div>
        <h3 style="margin: 8px 0 0 0; color: #0B2265; font-weight: 900; font-size: 1.2rem;">FSC AI AGENT GIO GIANG</h3>
        <span class="fpt-badge">V9.0 Multi-Unit</span>
    </div>
    """,
        unsafe_allow_html=True,
    )

    st.markdown(
        f"""
    <div style="background: #f8fafc; border: 1.5px solid #cbd5e1; border-radius: 8px; padding: 12px 14px; margin-bottom: 15px;">
        <p style="margin: 0; font-size: 11px; color: #64748b; font-weight: 700;">CÁN BỘ QA PHỤ TRÁCH:</p>
        <p style="margin: 3px 0 0 0; font-size: 14px; font-weight: 800; color: #0B2265;">👤 {st.session_state.config["owner"]}</p>
        <p style="margin: 2px 0 0 0; font-size: 12px; color: #0284c7; font-weight: 600;">🏢 {st.session_state.config["unit"]}</p>
        <div style="margin-top: 8px; display: flex; gap: 6px;">
            <span class="iso-badge" style="background: #0B2265; font-size: 11px; padding: 2px 8px;">ISO 21001:2018</span>
            <span class="iso-badge" style="background: #F26F21; font-size: 11px; padding: 2px 8px;">KWSR V9.0</span>
        </div>
    </div>
    """,
        unsafe_allow_html=True,
    )

    st.markdown("### ⚡ Thao Tác Nhanh")
    if st.button("🚀 Nạp Dữ Liệu Mẫu (01_Inputs/)", width="stretch"):
        sample_fsp = "01_Inputs/FSCHL_GGT5.xlsx"
        sample_swap = "01_Inputs/Doi tiet_Day thay.xlsx"
        if os.path.exists(sample_fsp):
            with st.spinner("Đang xử lý dữ liệu mẫu 7 bước..."):
                try:
                    res = run_fsc_pipeline(
                        sample_fsp,
                        sample_swap if os.path.exists(sample_swap) else None,
                    )
                    st.session_state.pipeline_results = res
                    st.success(
                        f"✅ Đã nạp thành công {len(res['df_table2'])} giáo viên!"
                    )
                    st.rerun()
                except (ValueError, KeyError, OSError, RuntimeError) as e:
                    st.error(f"Lỗi: {e}")
        else:
            st.error("Không tìm thấy file mẫu trong 01_Inputs/FSCHL_GGT5.xlsx")

    if st.button("🔄 Làm Mới Dữ Liệu", width="stretch"):
        st.session_state.pipeline_results = None
        st.session_state.execution_logs = []
        st.rerun()

    if st.button("🗑️ Xóa Lịch Sử Chat", width="stretch"):
        st.session_state.messages = [st.session_state.messages[0]]
        st.rerun()

    st.divider()
    if st.session_state.pipeline_results:
        res = st.session_state.pipeline_results
        over_q = len(res["df_table2"][res["df_table2"]["canh_bao_vuot_gio"].eq(True)])
        st.markdown(
            f"""
        <div style="font-size: 12.5px; color: #1e293b; line-height: 1.7; font-weight: 500;">
            <b style="color: #0B2265; font-size: 13px;">📊 THỐNG KÊ NHANH:</b><br>
            • Tổng GV: <b style="color: #0B2265;">{len(res["df_table2"])}</b> GV<br>
            • Tiết hợp lệ: <b style="color: #0284c7;">{len(res["df_table1"])}</b> lượt<br>
            • Tiết loại trừ: <b style="color: #e11d48;">{len(res["df_table3"])}</b> tiết<br>
            • Cảnh báo vượt 110h: <span style="color: {"#e11d48" if over_q > 0 else "#16a34a"}; font-weight: 800;">{over_q} GV</span><br>
            • Tổng quỹ lương: <b style="color: #F26F21;">{res["df_table2"]["Tổng Giờ Trả Lương"].sum():.2f}h</b>
        </div>
        """,
            unsafe_allow_html=True,
        )
    else:
        st.info("Chưa có dữ liệu được nạp vào hệ thống.")


# ==========================================
# 7. HEADER CHÍNH & TABS
# ==========================================
st.markdown(
    f"""
<div class="fpt-header">
    <div style="display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 10px;">
        <div>
            <div style="display: flex; align-items: center; gap: 12px;">
                <h1 style="margin: 0; font-size: 1.8rem; font-weight: 900; color: white !important; letter-spacing: -0.02em;">
                    FSC_AI AGENT GIO GIANG
                </h1>
                <span class="fpt-badge">Phiên Bản V9.0</span>
            </div>
            <p style="margin: 5px 0 0 0; font-size: 0.9rem; color: #bfdbfe; font-weight: 500;">
                🛡️ FPT QA Agent • Chuẩn hóa <b>Account GV</b> (Bỏ đuôi email) • Phân loại theo <b>CỘT LỚP</b> • Cô lập Quota 110h
            </p>
        </div>
        <div style="text-align: right;">
            <div class="iso-badge">✅ ISO 21001:2018 | KWSR</div>
            <p style="margin: 5px 0 0 0; font-size: 0.85rem; color: #fed7aa; font-weight: 700;">
                Owner: {st.session_state.config["owner"]}
            </p>
        </div>
    </div>
</div>
""",
    unsafe_allow_html=True,
)

tab_chat, tab_config, tab_pipeline, tab_dashboard, tab_tables = st.tabs(
    [
        "💬 1. Trợ Lý Hội Thoại QA",
        "⚙️ 2. Cấu Hình Hệ Số VÀ BỘ LỌC",
        "📁 3. Nạp Dữ Liệu & Pipeline",
        "📊 4. Dashboard TRỰC QUAN",
        "📋 5. Bảng Dữ Liệu 4 SHEET & Xuất File",
    ]
)

# ==========================================
# TAB 1: 💬 CHATBOT AI
# ==========================================
with tab_chat:
    st.markdown("### 💬 1. Trợ Lý Hội Thoại QA Giờ Giảng FSC (AI Agent)")
    st.caption(
        "Tra cứu giáo viên theo Account (ví dụ: hungnd, aint3), cảnh báo vượt định mức 110h, đối soát và điều khiển bằng ngôn ngữ tự nhiên."
    )

    st.markdown("**💡 Gợi ý câu hỏi nhanh:**")
    col_q1, col_q2, col_q3, col_q4, col_q5 = st.columns(5)
    quick_query = None
    if col_q1.button("🔍 Tra cứu Account hungnd", width="stretch"):
        quick_query = "Tra cứu giáo viên hungnd"
    if col_q2.button("🚨 GV vượt 110h", width="stretch"):
        quick_query = "Danh sách giáo viên cảnh báo vượt định mức 110h"
    if col_q3.button("📊 Tóm tắt quỹ giờ", width="stretch"):
        quick_query = "/report"
    if col_q4.button("🔄 Đối soát đổi tiết", width="stretch"):
        quick_query = "Kết quả đối soát lịch đổi tiết Bảng 4"
    if col_q5.button("🚫 Tiết loại trừ", width="stretch"):
        quick_query = "Thống kê các tiết bị loại trừ Bảng 3"

    for msg in st.session_state.messages:
        with st.chat_message(
            msg["role"], avatar="👤" if msg["role"] == "user" else "🤖"
        ):
            st.markdown(msg["content"])

    user_input = st.chat_input(
        "Nhập câu hỏi hoặc Slash Command (ví dụ: Tra cứu GV aint3, hungnd, /report)..."
    )
    prompt_to_process = quick_query or user_input
    if prompt_to_process:
        st.session_state.messages.append({"role": "user", "content": prompt_to_process})
        with st.chat_message("user", avatar="👤"):
            st.markdown(prompt_to_process)

        with (
            st.chat_message("assistant", avatar="🤖"),
            st.spinner("AI Agent đang tra cứu và phân tích dữ liệu..."),
        ):
            reply = answer_user_query(
                prompt_to_process,
                st.session_state.pipeline_results,
                st.session_state.config,
            )
            st.markdown(reply)
            st.session_state.messages.append({"role": "assistant", "content": reply})

# ==========================================
# TAB 2: ⚙️ CẤU HÌNH HỆ SỐ & BỘ LỌC
# ==========================================
with tab_config:
    st.markdown("### ⚙️ 2. Cấu Hình Hệ Số VÀ BỘ LỌC (Multi-Unit FSC Dynamic K)")
    st.markdown(
        """
    <div class="alert-info-box">
        <b>💡 Nguyên tắc phân loại:</b> Phân loại tiết dạy <b>HOÀN TOÀN DỰA VÀO CỘT LỚP</b>. Môn học <b>"Dự án PDP"</b> giảng dạy tại lớp chính khóa tính vào Giờ Chính Khóa. Account GV tự động bóc tách chuẩn xác bỏ đuôi email <code>@FPT.EDU.VN</code>.
    </div>
    """,
        unsafe_allow_html=True,
    )

    col_btn_preset, col_btn_reset, col_btn_json_exp, col_btn_json_imp = st.columns(
        [2, 2, 2, 2]
    )
    if col_btn_preset.button("🏢 Áp dụng Mẫu FSC HN", width="stretch"):
        st.session_state.config = DEFAULT_CONFIG.copy()
        st.success("Đã áp dụng mẫu chuẩn FSC Hà Nội!")
        st.rerun()

    if col_btn_reset.button("🔄 Đặt lại Mặc định", width="stretch"):
        st.session_state.config = DEFAULT_CONFIG.copy()
        st.success("Đã reset cấu hình về mặc định!")
        st.rerun()

    with col_btn_json_exp:
        json_cfg_str = json.dumps(st.session_state.config, indent=2, ensure_ascii=False)
        st.download_button(
            label="📥 Xuất Cấu Hình JSON",
            data=json_cfg_str,
            file_name=f"FSC_Config_GioGiang_{get_current_time_str('%Y%m%d')}.json",
            mime="application/json",
            width="stretch",
        )

    with col_btn_json_imp:
        uploaded_json_cfg = st.file_uploader(
            "Nhập JSON", type=["json"], label_visibility="collapsed"
        )
        if uploaded_json_cfg is not None:
            try:
                loaded_dict = json.load(uploaded_json_cfg)
                st.session_state.config.update(loaded_dict)
                st.success("Đã nạp cấu hình từ file JSON!")
                st.rerun()
            except json.JSONDecodeError as err:
                st.error(f"File JSON không hợp lệ: {err}")

    st.markdown("#### 1. Hệ số Giờ Dạy & Định Mức Chuẩn")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.session_state.config["k_th"] = st.number_input(
            "Hệ số Tiểu Học (K_TH) [35p/tiết]",
            value=float(st.session_state.config["k_th"]),
            step=0.000000001,
            format="%.9f",
        )
        st.session_state.config["k_hsg"] = st.number_input(
            "Hệ số Bồi Dưỡng HSG (K_HSG)",
            value=float(st.session_state.config["k_hsg"]),
            step=0.05,
            format="%.2f",
        )
        st.session_state.config["quota_threshold"] = st.number_input(
            "Ngưỡng Cảnh Báo Vượt Định Mức (Giờ)",
            value=float(st.session_state.config["quota_threshold"]),
            step=5.0,
            format="%.1f",
        )

    with c2:
        st.session_state.config["k_thcs"] = st.number_input(
            "Hệ số THCS/THPT (K_THCS) [45p/tiết]",
            value=float(st.session_state.config["k_thcs"]),
            step=0.01,
            format="%.4f",
        )
        st.session_state.config["k_pd"] = st.number_input(
            "Hệ số Phụ Đạo / Đội Tuyển (K_PD)",
            value=float(st.session_state.config["k_pd"]),
            step=0.05,
            format="%.2f",
        )

    with c3:
        st.session_state.config["k_clb"] = st.number_input(
            "Hệ số Câu Lạc Bộ (K_CLB)",
            value=float(st.session_state.config["k_clb"]),
            step=0.05,
            format="%.2f",
        )
        st.session_state.config["k_sukien"] = st.number_input(
            "Hệ số Sự Kiện / Dự Án Khác (K_SK)",
            value=float(st.session_state.config["k_sukien"]),
            step=0.05,
            format="%.2f",
        )

    st.markdown("#### 2. Danh Mục Từ Khóa Loại Trừ (Bảng 3)")
    current_exclusions_str = ", ".join(st.session_state.config["exclusions"])
    new_exclusions_str = st.text_area(
        "Nhập các từ khóa loại trừ (phân cách bằng dấu phẩy):",
        value=current_exclusions_str,
    )
    st.session_state.config["exclusions"] = [
        k.strip().upper() for k in new_exclusions_str.split(",") if k.strip()
    ]

    st.markdown("#### 3. Biểu Thức Regex Nhận Diện Tiết Đặc Thù Cơ Bản (Cột Lớp)")
    r1, r2 = st.columns(2)
    with r1:
        st.session_state.config["regex_hsg_lop"] = st.text_input(
            "Regex Lớp HSG:", value=st.session_state.config["regex_hsg_lop"]
        )
        st.session_state.config["regex_pd_lop"] = st.text_input(
            "Regex Lớp Phụ Đạo / Đội Tuyển:",
            value=st.session_state.config["regex_pd_lop"],
        )
    with r2:
        st.session_state.config["regex_clb_lop"] = st.text_input(
            "Regex Lớp CLB:", value=st.session_state.config["regex_clb_lop"]
        )
        st.session_state.config["regex_sukien_lop"] = st.text_input(
            "Regex Sự Kiện:", value=st.session_state.config["regex_sukien_lop"]
        )

    st.markdown(
        "#### 4. 🏢 Danh Mục Tiết Dạy Tùy Chỉnh Theo Đơn Vị (Custom Categories)"
    )
    st.caption(
        "Thêm các loại tiết riêng của trường bạn nếu không thuộc các nhóm cơ bản ở trên."
    )

    # Form thêm loại tiết tùy chỉnh mới
    with st.expander("➕ Thêm Loại Tiết Tùy Chỉnh Mới", expanded=False):
        c_cat1, c_cat2 = st.columns(2)
        with c_cat1:
            new_cat_name = st.text_input(
                "Tên Loại Tiết Mới (Label):",
                placeholder="Ví dụ: STEM Đặc Thù, Dự Án Trải Nghiệm...",
                key="input_new_cat_name",
            )
            new_cat_k = st.number_input(
                "Hệ số Quy Đổi (K):",
                value=1.00,
                step=0.05,
                format="%.2f",
                key="input_new_cat_k",
            )
        with c_cat2:
            new_cat_regex = st.text_input(
                "Regex Nhận Diện Trên Cột Lớp:",
                placeholder="Ví dụ: STEM|TRẢI NGHIỆM",
                key="input_new_cat_regex",
            )
            new_cat_quota = st.checkbox(
                "Tính vào Quota 110h (Giờ Xét Định Mức)",
                value=False,
                key="input_new_cat_quota",
            )

        if st.button("💾 Lưu Loại Tiết Tùy Chỉnh", key="btn_save_custom_cat"):
            if not new_cat_name.strip() or not new_cat_regex.strip():
                st.error("Vui lòng nhập đầy đủ Tên loại tiết và Biểu thức Regex!")
            else:
                st.session_state.config["custom_categories"].append(
                    {
                        "label": new_cat_name.strip(),
                        "k": float(new_cat_k),
                        "regex": new_cat_regex.strip().upper(),
                        "quota": bool(new_cat_quota),
                    }
                )
                st.success(f"✅ Đã thêm loại tiết: {new_cat_name} (K = {new_cat_k})")
                st.rerun()

    # Hiển thị danh sách các custom categories
    custom_cats = st.session_state.config.get("custom_categories", [])
    if not custom_cats:
        st.info(
            "Chưa có loại tiết tùy chỉnh bổ sung. Quý Thầy/Cô có thể nhấn '➕ Thêm Loại Tiết Tùy Chỉnh Mới' ở trên để bổ sung."
        )
    else:
        st.markdown("**Danh sách loại tiết tùy chỉnh hiện có:**")
        num_cols = min(len(custom_cats), 3)
        cols_cats = st.columns(num_cols)
        for idx, cat in enumerate(custom_cats):
            col_target = cols_cats[idx % num_cols]
            with col_target:
                quota_str = "🟢 CÓ TÍNH" if cat.get("quota") else "🔴 CÔ LẬP (KHÔNG)"
                st.markdown(
                    f"""
                <div style="background: white; border: 1.5px solid #10b981; border-radius: 8px; padding: 10px 12px; margin-bottom: 8px; border-left: 4px solid #F26F21; background-color: #f0fdf4;">
                    <div style="font-weight: 800; font-size: 13.5px; color: #0B2265;">🏷️ {cat.get("label", "Tùy chỉnh")}</div>
                    <div style="font-size: 11.5px; color: #475569; margin-top: 4px; line-height: 1.6;">
                        • Hệ số K: <b style="color: #F26F21;">{cat.get("k", 1.0)}</b><br>
                        • Regex: <code style="font-size: 10.5px;">{cat.get("regex", "")}</code><br>
                        • Quota 110h: <b style="color: #0B2265;">{quota_str}</b>
                    </div>
                </div>
                """,
                    unsafe_allow_html=True,
                )
                if st.button(
                    f"🗑️ Xóa '{cat.get('label')}'",
                    key=f"del_cat_{idx}",
                    width="stretch",
                ):
                    st.session_state.config["custom_categories"].pop(idx)
                    st.rerun()

# ==========================================
# TAB 3: 📁 NẠP DỮ LIỆU & PIPELINE
# ==========================================
with tab_pipeline:
    st.markdown("### 📁 3. Nạp Dữ Liệu & Pipeline 7 Bước")
    col_up1, col_up2 = st.columns(2)
    with col_up1:
        uploaded_fsp = st.file_uploader(
            "📥 1. File FSP Giờ Giảng (.xlsx) [Bắt buộc]:",
            type=["xlsx", "xls"],
        )
    with col_up2:
        uploaded_swap = st.file_uploader(
            "📥 2. File Đổi Tiết / Dạy Thay (.xlsx) [Tùy chọn]:",
            type=["xlsx", "xls"],
        )

    col_act1, col_act2, _ = st.columns([3, 3, 4])
    run_clicked = col_act1.button("⚡ KHỞI CHẠY TIẾN TRÌNH XỬ LÝ", width="stretch")
    load_sample_clicked = col_act2.button(
        "🚀 DÙNG FILE MẪU CÓ SẴN (01_Inputs/)", width="stretch"
    )

    if run_clicked:
        if not uploaded_fsp:
            st.error("Vui lòng tải lên file FSP Giờ Giảng trước khi bấm xử lý!")
        else:
            with st.spinner("Đang thực hiện quy trình 7 bước..."):
                try:
                    res = run_fsc_pipeline(
                        uploaded_fsp, uploaded_swap, st.session_state.config
                    )
                    st.session_state.pipeline_results = res
                    st.success("🎉 TIẾN TRÌNH XỬ LÝ V9.0 ĐÃ HOÀN TẤT THÀNH CÔNG!")
                except (ValueError, KeyError, OSError, RuntimeError) as e:
                    st.error(f"LỖI: {e}")

    if load_sample_clicked:
        sample_fsp = "01_Inputs/FSCHL_GGT5.xlsx"
        sample_swap = "01_Inputs/Doi tiet_Day thay.xlsx"
        if os.path.exists(sample_fsp):
            with st.spinner("Đang nạp và xử lý dữ liệu mẫu 01_Inputs/..."):
                try:
                    res = run_fsc_pipeline(
                        sample_fsp,
                        sample_swap if os.path.exists(sample_swap) else None,
                        st.session_state.config,
                    )
                    st.session_state.pipeline_results = res
                    st.success("🎉 ĐÃ XỬ LÝ XONG FILE MẪU THÀNH CÔNG!")
                except (ValueError, KeyError, OSError, RuntimeError) as e:
                    st.error(f"LỖI: {e}")
        else:
            st.error("Không tìm thấy file mẫu trong 01_Inputs/FSCHL_GGT5.xlsx")

    st.markdown("#### 🧭 Quy Trình Xử Lý Chuẩn 7 Bước (KWSR Pipeline)")
    steps = [
        (
            "1",
            "Nạp file FSP & Tự động phát hiện Smart Header",
            "Hoàn thành" if st.session_state.pipeline_results else "Chờ",
        ),
        (
            "2",
            "Bóc tách chuỗi lớp ghép & Chuẩn hóa Account GV",
            "Hoàn thành" if st.session_state.pipeline_results else "Chờ",
        ),
        (
            "3",
            "Lọc loại trừ từ khóa không tính giờ sang Bảng 3",
            "Hoàn thành" if st.session_state.pipeline_results else "Chờ",
        ),
        (
            "4",
            "Phân loại tiết dạy theo Cột Lớp & Custom Categories sang Bảng 1",
            "Hoàn thành" if st.session_state.pipeline_results else "Chờ",
        ),
        (
            "5",
            "Đối soát tọa độ 3 chiều [Ngày + Tiết + Lớp] sang Bảng 4",
            "Hoàn thành" if st.session_state.pipeline_results else "Chờ",
        ),
        (
            "6",
            "Tổng hợp Số Tiết & Số Giờ, cô lập Quota 110h sang Bảng 2",
            "Hoàn thành" if st.session_state.pipeline_results else "Chờ",
        ),
        (
            "7",
            "Xuất file Excel 4 Sheets chuẩn đẹp 03_Outputs/",
            "Hoàn thành" if st.session_state.pipeline_results else "Chờ",
        ),
    ]
    cols_s1, cols_s2 = st.columns(2)
    for i, (num, title, status) in enumerate(steps):
        target_col = cols_s1 if i < 4 else cols_s2
        with target_col:
            badge_color = "#16a34a" if status == "Hoàn thành" else "#64748b"
            bg_badge = "#f0fdf4" if status == "Hoàn thành" else "#f1f5f9"
            st.markdown(
                f"""
            <div style="background: white; border: 1.5px solid #cbd5e1; border-radius: 8px; padding: 10px 14px; margin-bottom: 8px; display: flex; align-items: center; justify-content: space-between;">
                <div>
                    <span class="step-badge">{num}</span>
                    <span style="font-size: 13.5px; font-weight: 700; color: #0B2265;">{title}</span>
                </div>
                <span style="font-size: 11.5px; font-weight: 800; color: {badge_color}; background: {bg_badge}; padding: 3px 10px; border-radius: 9999px;">{status}</span>
            </div>
            """,
                unsafe_allow_html=True,
            )

    if st.session_state.execution_logs:
        st.markdown("#### 📜 Nhật Ký Tiến Trình (Live Logs)")
        st.code("\n".join(st.session_state.execution_logs), language="bash")

# ==========================================
# TAB 4: 📊 DASHBOARD TRỰC QUAN
# ==========================================
with tab_dashboard:
    st.markdown("### 📊 4. Dashboard TRỰC QUAN Giờ Giảng FSC")
    if not st.session_state.pipeline_results:
        st.warning(
            "Chưa có dữ liệu kết quả. Hãy nạp file tại Tab 3 hoặc bấm 'Nạp Dữ Liệu Mẫu'."
        )
    else:
        res = st.session_state.pipeline_results
        df_t1 = res["df_table1"]
        df_t2 = res["df_table2"]
        df_t3 = res["df_table3"]
        df_t4 = res["df_table4"]

        over_q_list = df_t2[df_t2["canh_bao_vuot_gio"].eq(True)]
        sum_th = df_t2["Số Giờ TH"].sum()
        sum_thcs = df_t2["Số Giờ THCS/THPT"].sum()
        sum_hsg = df_t2["Số Giờ HSG"].sum()
        sum_pd = df_t2["Số Giờ PD_ĐT"].sum()
        sum_clb = df_t2["Số Giờ CLB"].sum()
        sum_sk = df_t2["Số Giờ Sự Kiện"].sum()
        sum_cust = (
            df_t2["Số Giờ Tùy Chỉnh"].sum()
            if "Số Giờ Tùy Chỉnh" in df_t2.columns
            else 0.0
        )
        sum_total_salary = df_t2["Tổng Giờ Trả Lương"].sum()
        sum_total_quota = df_t2["Tổng Giờ Xét Định Mức"].sum()

        m1, m2, m3, m4, m5 = st.columns(5)
        with m1:
            st.markdown(
                f"""
            <div class="metric-card">
                <div class="metric-label">Tổng Giáo Viên</div>
                <div class="metric-val">{len(df_t2)}</div>
                <div style="font-size: 11.5px; color: #16a34a; font-weight: 700;">👥 Bảng 2 Tổng hợp</div>
            </div>
            """,
                unsafe_allow_html=True,
            )
        with m2:
            st.markdown(
                f"""
            <div class="metric-card">
                <div class="metric-label">Tiết Hợp Lệ</div>
                <div class="metric-val">{len(df_t1)}</div>
                <div style="font-size: 11.5px; color: #0284c7; font-weight: 700;">📚 Bảng 1 Chi tiết</div>
            </div>
            """,
                unsafe_allow_html=True,
            )
        with m3:
            st.markdown(
                f"""
            <div class="metric-card">
                <div class="metric-label">Tiết Loại Trừ</div>
                <div class="metric-val">{len(df_t3)}</div>
                <div style="font-size: 11.5px; color: #e11d48; font-weight: 700;">🚫 Bảng 3 SHL/Tự học</div>
            </div>
            """,
                unsafe_allow_html=True,
            )
        with m4:
            top_border_c = "#e11d48" if len(over_q_list) > 0 else "#16a34a"
            val_c = "#e11d48" if len(over_q_list) > 0 else "#16a34a"
            st.markdown(
                f"""
            <div class="metric-card" style="border-top-color: {top_border_c};">
                <div class="metric-label">Vượt Định Mức</div>
                <div class="metric-val" style="color: {val_c};">{len(over_q_list)}</div>
                <div style="font-size: 11.5px; color: #64748b; font-weight: 700;">⚠️ > 110 giờ chính khóa</div>
            </div>
            """,
                unsafe_allow_html=True,
            )
        with m5:
            st.markdown(
                f"""
            <div class="metric-card">
                <div class="metric-label">Tổng Quỹ Giờ Lương</div>
                <div class="metric-val" style="color: #F26F21;">{sum_total_salary:.1f}h</div>
                <div style="font-size: 11.5px; color: #ea580c; font-weight: 700;">💰 Định mức: {sum_total_quota:.1f}h</div>
            </div>
            """,
                unsafe_allow_html=True,
            )

        st.markdown("---")
        chart_col1, chart_col2 = st.columns(2)
        with chart_col1:
            st.markdown("#### 🍩 Cơ Cấu Phân Bổ Số Giờ Theo Loại Tiết")
            pie_data = pd.DataFrame(
                {
                    "Loại Tiết": [
                        "Chính khóa TH",
                        "Chính khóa THCS/THPT",
                        "HSG",
                        "Phụ đạo/ĐT",
                        "CLB",
                        "Sự kiện/Khác",
                        "Tùy chỉnh riêng",
                    ],
                    "Số Giờ Quy Đổi": [
                        sum_th,
                        sum_thcs,
                        sum_hsg,
                        sum_pd,
                        sum_clb,
                        sum_sk,
                        sum_cust,
                    ],
                }
            )
            pie_data = pie_data[pie_data["Số Giờ Quy Đổi"] > 0]
            fig_pie = px.pie(
                pie_data,
                values="Số Giờ Quy Đổi",
                names="Loại Tiết",
                hole=0.45,
                color_discrete_sequence=[
                    "#f59e0b",
                    "#3b82f6",
                    "#eab308",
                    "#a855f7",
                    "#10b981",
                    "#f43f5e",
                    "#6366f1",
                ],
            )
            fig_pie.update_layout(
                margin={"t": 20, "b": 20, "l": 20, "r": 20},
                legend={"orientation": "h", "y": -0.1},
            )
            st.plotly_chart(fig_pie, width="stretch")

        with chart_col2:
            st.markdown("#### 📊 Top 10 Giáo Viên Có Quỹ Giờ Giảng Cao Nhất")
            top_gv = df_t2.head(10).copy()
            top_gv["Giờ Đặc Thù"] = (
                top_gv["Tổng Giờ Trả Lương"] - top_gv["Tổng Giờ Xét Định Mức"]
            )

            fig_bar = go.Figure(
                data=[
                    go.Bar(
                        name="Giờ Xét Định Mức (Chính Khóa)",
                        x=top_gv["Account Giáo Viên"],
                        y=top_gv["Tổng Giờ Xét Định Mức"],
                        marker_color="#0B2265",
                    ),
                    go.Bar(
                        name="Giờ Đặc Thù / Khác",
                        x=top_gv["Account Giáo Viên"],
                        y=top_gv["Giờ Đặc Thù"],
                        marker_color="#F26F21",
                    ),
                ]
            )
            fig_bar.update_layout(
                barmode="stack",
                margin={"t": 20, "b": 20, "l": 20, "r": 20},
                legend={"orientation": "h", "y": 1.1},
                xaxis={"tickangle": -45},
            )
            st.plotly_chart(fig_bar, width="stretch")

# ==========================================
# TAB 5: 📋 BẢNG DỮ LIỆU 4 SHEET & XUẤT FILE
# ==========================================
with tab_tables:
    st.markdown("### 📋 5. Bảng Dữ Liệu 4 SHEET & Xuất File Kết Quả")
    st.markdown(
        """
    <div class="table-highlight-box">
        <b style="color: #0B2265; font-size: 13.5px;">🌿 Định Dạng Bảng Dữ Liệu:</b> Cột <b>Số Tiết</b> (Xanh dương nhạt), Cột <b>Số Giờ</b> (Xanh lá cây), Cột <b>Tổng Giờ & Cảnh Báo</b> (Cam nhạt), Tiêu đề cột (Xanh dương đậm 12px, béo).
    </div>
    """,
        unsafe_allow_html=True,
    )

    if not st.session_state.pipeline_results:
        st.warning("Chưa có dữ liệu kết quả để hiển thị.")
    else:
        res = st.session_state.pipeline_results
        c_exp1, c_exp2, _ = st.columns([3, 3, 4])
        with c_exp1:
            file_ts = get_current_time_str("%Y-%m-%d")
            st.download_button(
                label="📥 Tải File Excel 4 Sheets (.xlsx)",
                data=res["excel_bytes"],
                file_name=f"{file_ts}_FPT_QA_BaoCao_GioGiang_00234640_Pham_Thi_Minh_Hai.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )

        today_str = get_current_time_str("%Y-%m-%d")
        now_str = get_current_time_str("%Y-%m-%d %H:%M:%S")
        df_t2 = res["df_table2"]
        over_q_list = df_t2[df_t2["canh_bao_vuot_gio"].eq(True)]

        md_report = f"""# BÁO CÁO TỔNG HỢP KIỂM SOÁT VÀ ĐỐI SOÁT GIỜ GIẢNG (V9.0)
- **Hệ thống thực hiện:** FSC_AI AGENT GIO GIANG (FPT QA Agent V9.0)
- **Cán bộ QA phụ trách:** {st.session_state.config["owner"]} ({st.session_state.config["unit"]})
- **Thời gian xuất:** {now_str}
- **Tiêu chuẩn áp dụng:** ISO 21001:2018 | KWSR QA Pipeline
- **Tổng số giáo viên:** {len(df_t2)} GV
- **Tổng Quỹ Giờ Trả Lương:** {df_t2["Tổng Giờ Trả Lương"].sum():.2f} giờ
- **Tổng Quỹ Giờ Xét Định Mức:** {df_t2["Tổng Giờ Xét Định Mức"].sum():.2f} giờ
- **Số GV cảnh báo vượt định mức (>110h):** {len(over_q_list)} GV
"""
        with c_exp2:
            st.download_button(
                label="📄 Tải Báo Cáo Tổng Hợp (.md)",
                data=md_report,
                file_name=f"{today_str}_BaoCao_TongHop_00234640_Pham_Thi_Minh_Hai.md",
                mime="text/markdown",
                width="stretch",
            )

        st.markdown("---")
        filter_query = (
            st.text_input("🔍 Tìm kiếm nhanh theo Account GV, Lớp, Môn học...", "")
            .strip()
            .upper()
        )

        # Tab phân loại 4 bảng có Font chữ MÀU XANH DƯƠNG, BÉO, CỠ CHỮ 14px
        sub_tab1, sub_tab2, sub_tab3, sub_tab4 = st.tabs(
            [
                f"Bảng 1: Chi Tiết Tiết Dạy ({len(res['df_table1'])})",
                f"Bảng 2: Tổng Hợp Số Tiết & Số Giờ Theo GV ({len(res['df_table2'])})",
                f"Bảng 3: Tiết Loại Trừ ({len(res['df_table3'])})",
                f"Bảng 4: Log Đối Soát Đổi Tiết ({len(res['df_table4'])})",
            ]
        )

        with sub_tab1:
            df1_show = res["df_table1"].copy()
            if "Custom_Quota_Eligible" in df1_show.columns:
                df1_show = df1_show.drop(columns=["Custom_Quota_Eligible"])
            if filter_query:
                df1_show = df1_show[
                    df1_show["Account GV"]
                    .str.upper()
                    .str.contains(filter_query, na=False)
                    | df1_show["Lớp"].str.upper().str.contains(filter_query, na=False)
                    | df1_show["Môn học"]
                    .str.upper()
                    .str.contains(filter_query, na=False)
                ]
            # Hiển thị bảng với màu sắc cột chuyên biệt và số chuẩn sạch (2 chữ số thập phân cho tiết THCS/THPT)
            st.dataframe(style_fsc_dataframe(df1_show, is_table1=True), width="stretch")

        with sub_tab2:
            df2_show = res["df_table2"].copy()
            if filter_query:
                df2_show = df2_show[
                    df2_show["Account Giáo Viên"]
                    .str.upper()
                    .str.contains(filter_query, na=False)
                ]

            show_only_over = st.checkbox(
                "🚨 Chỉ hiển thị giáo viên vượt định mức (>110h)", value=False
            )
            if show_only_over:
                df2_show = df2_show[df2_show["canh_bao_vuot_gio"].eq(True)]

            df2_final = (
                df2_show.drop(columns=["canh_bao_vuot_gio"])
                if "canh_bao_vuot_gio" in df2_show.columns
                else df2_show
            )
            # Hiển thị bảng với màu sắc cột chuyên biệt và số chuẩn sạch
            st.dataframe(style_fsc_dataframe(df2_final), width="stretch")

        with sub_tab3:
            df3_show = res["df_table3"].copy()
            if filter_query:
                df3_show = df3_show[
                    df3_show["Account GV"]
                    .str.upper()
                    .str.contains(filter_query, na=False)
                    | df3_show["Lớp"].str.upper().str.contains(filter_query, na=False)
                    | df3_show["Môn học"]
                    .str.upper()
                    .str.contains(filter_query, na=False)
                ]
            st.dataframe(style_fsc_dataframe(df3_show), width="stretch")

        with sub_tab4:
            df4_show = res["df_table4"].copy()
            if filter_query:
                df4_show = df4_show[
                    df4_show["GV Theo TKB"]
                    .str.upper()
                    .str.contains(filter_query, na=False)
                    | df4_show["GV Dạy Thay"]
                    .str.upper()
                    .str.contains(filter_query, na=False)
                    | df4_show["Lớp"].str.upper().str.contains(filter_query, na=False)
                ]
            st.dataframe(style_fsc_dataframe(df4_show), width="stretch")

# ==========================================
# 8. FOOTER
# ==========================================
st.markdown("---")
st.markdown(
    f"""
<div style="text-align: center; font-size: 12px; color: #94a3b8; padding: 10px 0;">
    <b>FSC_AI AGENT GIO GIANG (V9.0)</b> — FPT Education QA-KSCL<br>
    Hệ thống chuẩn hóa & đối soát giờ giảng liên cấp • Cán bộ phụ trách: <b>{st.session_state.config["owner"]}</b> • Tiêu chuẩn ISO 21001:2018
</div>
""",
    unsafe_allow_html=True,
)
